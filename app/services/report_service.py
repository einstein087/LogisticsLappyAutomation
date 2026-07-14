"""
Weekly compliance report generator.

Produces an Excel workbook (.xlsx) summarising all laptop-provisioning
requests for a given period and, optionally, uploads it to SharePoint.
"""

import logging
import os
import tempfile
from datetime import datetime, timedelta, timezone

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from flask import current_app

from app.services import sharepoint_service

logger = logging.getLogger(__name__)

_HEADERS = [
    "Request ID", "Employee Name", "Employee Email", "Department",
    "Manager", "Manager Email", "Laptop Model", "Asset Tag",
    "Status", "Priority", "Requested On", "Approved On", "Dispatched On",
    "Delivered On", "Notes",
]

_STATUS_FILL = {
    "pending":    "FFF3CD",
    "approved":   "D4EDDA",
    "rejected":   "F8D7DA",
    "in_progress":"CCE5FF",
    "dispatched": "D6EAF8",
    "delivered":  "C3E6CB",
    "cancelled":  "E2E3E5",
}


def _cell_value(req, field: str) -> str:
    val = getattr(req, field, None)
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d %H:%M")
    return str(val)


def generate_weekly_report(weeks_back: int = 0) -> str:
    """
    Build a .xlsx compliance report for the completed calendar week
    (or *weeks_back* weeks earlier) and return the local file path.
    """
    from app.models import LaptopRequest  # imported here to avoid circular import

    now = datetime.now(timezone.utc)
    # ISO week boundaries
    week_start = now - timedelta(days=now.weekday() + 7 * (1 + weeks_back))
    week_start = week_start.replace(hour=0, minute=0, second=0, microsecond=0)
    week_end = week_start + timedelta(days=7)

    requests = (
        LaptopRequest.query
        .filter(LaptopRequest.created_at >= week_start)
        .filter(LaptopRequest.created_at < week_end)
        .order_by(LaptopRequest.created_at)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Provisioning Report"

    # --- Title row ---
    ws.merge_cells("A1:O1")
    title_cell = ws["A1"]
    title_cell.value = (
        f"Laptop Provisioning Weekly Report – "
        f"W/C {week_start.strftime('%d %b %Y')}"
    )
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # --- Summary row ---
    ws.append([])  # blank
    total = len(requests)
    approved = sum(1 for r in requests if r.status in ("approved", "in_progress", "dispatched", "delivered"))
    rejected = sum(1 for r in requests if r.status == "rejected")
    pending = sum(1 for r in requests if r.status == "pending")
    ws.append([
        f"Total: {total}",
        f"Approved: {approved}",
        f"Rejected: {rejected}",
        f"Pending: {pending}",
    ])

    ws.append([])  # blank

    # --- Header row ---
    header_row = ws.max_row + 1
    ws.append(_HEADERS)
    header_fill = PatternFill("solid", fgColor="343A40")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, _ in enumerate(_HEADERS, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # --- Data rows ---
    _fields = [
        "id", "employee_name", "employee_email", "department",
        "manager_name", "manager_email", "laptop_model", "asset_tag",
        "status", "priority", "created_at", "approved_at", "dispatched_at",
        "delivered_at", "notes",
    ]
    for req in requests:
        row = [_cell_value(req, f) for f in _fields]
        ws.append(row)
        # Colour-code by status
        fill_colour = _STATUS_FILL.get(req.status, "FFFFFF")
        row_idx = ws.max_row
        for col_idx in range(1, len(_HEADERS) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor=fill_colour)

    # --- Column widths ---
    col_widths = [10, 22, 28, 18, 20, 28, 20, 14, 14, 10, 18, 18, 18, 18, 30]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # --- Save ---
    reports_dir = os.path.join(current_app.root_path, "..", "reports")
    os.makedirs(reports_dir, exist_ok=True)
    filename = f"provisioning_report_{week_start.strftime('%Y_W%W')}.xlsx"
    file_path = os.path.join(reports_dir, filename)
    wb.save(file_path)
    logger.info("Weekly report saved: %s", file_path)
    return file_path


def generate_and_upload() -> dict:
    """Generate this week's report and upload it to SharePoint."""
    file_path = generate_weekly_report()
    filename = os.path.basename(file_path)
    web_url = sharepoint_service.upload_report(current_app.config, file_path, filename)
    return {"file_path": file_path, "filename": filename, "sharepoint_url": web_url}
